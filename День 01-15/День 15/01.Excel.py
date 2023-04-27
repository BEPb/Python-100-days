"""
Python 3.10 создание Excel документа
Название файла '01.Excel.py'

Version: 0.1
Author: Andrej Marinchenko
Date: 2023-04-27
"""
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime

workbook = Workbook()  # создаем экземпляр класса
sheet = workbook.active  # создадим лист

# создадим список данных
data = [
    [1001, '323', '33333', '13123456789'],
    [1002, '564', '55543', '13233445566']
]

sheet.append(['12', '13', '14', '15'])  # запишем в наш лист следующие данные

'''далее в ранее созданный лист добавим данные из нашего списка'''
for row in data:  # циклическое считывание строки из списка данных
    sheet.append(row)  # добавляем строку в лист


tab = Table(displayName="Table1", ref="A1:E5")  # создадим экземпляр таблицы

# установим стиль таблицы
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)

# в созданный лист добавляем нашу таблицу
sheet.add_table(tab)
workbook.save('./res/97.xlsx')  # сохраним результат в файл

wb = Workbook()  # создаем экземпляр класса
ws = wb.active  # создадим лист

ws['A1'] = 42  # запишем в ячейку А1 значение 42
ws.append([1, 2, 3])  # добавим в таблицу следующий список, запись будет осуществлена в следующую строку (т.е. А2 - С2)
ws['A2'] = datetime.datetime.now()  # перезапишем значение ячийки А2, указав в ней текущее время

wb.save("./res/sample.xlsx") # сохраним результат в файл
