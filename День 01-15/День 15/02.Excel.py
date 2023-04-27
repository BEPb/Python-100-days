"""
Python 3.10 чтение Excel документа
Название файла '02.Excel.py'

Version: 0.1
Author: Andrej Marinchenko
Date: 2023-04-27
"""

from openpyxl import load_workbook  # импорт библиотеки для чтения

workbook = load_workbook('./res/99.xlsx')  # считываем файл
print(workbook.sheetnames)  # печать названий всех листов файла (их два ['название листа файла эксель', 'лист номер 2'])
sheet = workbook[workbook.sheetnames[0]]  # присвоим переменной название первого листа
print(sheet.title)  # печать названия первого листа

for row in range(2, 7):  # используем цикл для прохода по строкам со 2-й по 6-ю
    for col in range(65, 70):   # используем цикл для прохода по колонкам с 65-й по 70-ю
        cell_index = chr(col) + str(row)  # получаем индекс ячейки номер колонки + номер строки
        print(sheet[cell_index].value, end='\t')  # печатаем все содержимые ячейки в одной строке
    print()  # переход на новую строку
